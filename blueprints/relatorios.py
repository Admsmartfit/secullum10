from datetime import datetime, date
from io import BytesIO
from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from extensions import db
from models import Batida, Funcionario

relatorios_bp = Blueprint('relatorios', __name__)

DIAS_PT = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']


def _query_batidas(data_inicio, data_fim, dept=None, func_id=None):
    q = (
        Batida.query
        .filter(Batida.data >= data_inicio, Batida.data <= data_fim)
        .join(Funcionario)
        .filter(Funcionario.ativo == True)
    )
    if dept:
        q = q.filter(Funcionario.departamento == dept)
    if func_id:
        q = q.filter(Batida.funcionario_id == func_id)
    return q.order_by(Funcionario.nome, Batida.data, Batida.hora)


@relatorios_bp.route('/relatorios')
@login_required
def relatorios():
    data_inicio_str = request.args.get('data_inicio', date.today().strftime('%Y-%m-%d'))
    data_fim_str    = request.args.get('data_fim',    date.today().strftime('%Y-%m-%d'))
    dept_sel  = request.args.get('dept', '')
    func_sel  = request.args.get('func_id', '')

    data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
    data_fim    = datetime.strptime(data_fim_str,    '%Y-%m-%d').date()

    batidas_query = _query_batidas(data_inicio, data_fim, dept_sel or None, func_sel or None).all()

    total_batidas = len(batidas_query)
    funcionarios_unicos = len(set(b.funcionario_id for b in batidas_query))
    inconsistencias = sum(1 for b in batidas_query if b.inconsistente)

    por_departamento = {}
    por_funcionario  = {}
    for b in batidas_query:
        dept = b.funcionario.departamento or 'Sem Departamento'
        por_departamento[dept] = por_departamento.get(dept, 0) + 1
        fid = b.funcionario_id
        if fid not in por_funcionario:
            por_funcionario[fid] = {'nome': b.funcionario.nome, 'batidas': 0}
        por_funcionario[fid]['batidas'] += 1

    ranking = sorted(por_funcionario.values(), key=lambda x: x['batidas'], reverse=True)[:10]

    # Para dropdowns de filtro
    departamentos = [r[0] for r in
        db.session.query(Funcionario.departamento)
        .filter(Funcionario.ativo == True, Funcionario.departamento.isnot(None))
        .distinct().order_by(Funcionario.departamento).all()]
    funcionarios = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()

    return render_template(
        'relatorios.html',
        data_inicio=data_inicio_str,
        data_fim=data_fim_str,
        dept_sel=dept_sel,
        func_sel=func_sel,
        total_batidas=total_batidas,
        funcionarios_unicos=funcionarios_unicos,
        inconsistencias=inconsistencias,
        por_departamento=por_departamento,
        ranking=ranking,
        departamentos=departamentos,
        funcionarios=funcionarios,
    )


@relatorios_bp.route('/relatorios/exportar-pontos')
@login_required
def exportar_pontos():
    """Exporta todas as batidas do período em Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data_inicio_str = request.args.get('data_inicio', date.today().strftime('%Y-%m-%d'))
    data_fim_str    = request.args.get('data_fim',    date.today().strftime('%Y-%m-%d'))
    dept_sel  = request.args.get('dept', '') or None
    func_id   = request.args.get('func_id', '') or None

    data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
    data_fim    = datetime.strptime(data_fim_str,    '%Y-%m-%d').date()

    batidas = _query_batidas(data_inicio, data_fim, dept_sel, func_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pontos'

    # Estilos
    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_fill  = PatternFill('solid', fgColor='1E293B')
    center    = Alignment(horizontal='center', vertical='center')
    thin      = Side(style='thin', color='CBD5E1')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    warn_fill = PatternFill('solid', fgColor='FEE2E2')

    headers = [
        'Funcionário', 'ID', 'CPF', 'Departamento', 'Função',
        'Data', 'Dia da Semana', 'Hora', 'Tipo', 'Origem', 'Inconsistente',
    ]
    col_widths = [30, 14, 16, 24, 20, 12, 14, 10, 12, 14, 14]

    # Cabeçalho
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    # Dados
    for row_idx, b in enumerate(batidas, start=2):
        f = b.funcionario
        dia_str = DIAS_PT[b.data.weekday()] if b.data else ''
        row_data = [
            f.nome,
            f.id,
            f.cpf or '',
            f.departamento or '',
            f.funcao or '',
            b.data.strftime('%d/%m/%Y') if b.data else '',
            dia_str,
            b.hora or '',
            b.tipo or '',
            b.origem or '',
            'Sim' if b.inconsistente else 'Não',
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if b.inconsistente:
                cell.fill = warn_fill

    # Rodapé com totais
    last = len(batidas) + 2
    ws.cell(row=last, column=1, value=f'Total: {len(batidas)} batidas').font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'pontos_{data_inicio_str}_a_{data_fim_str}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
