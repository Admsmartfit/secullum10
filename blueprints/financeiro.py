from datetime import date, datetime, timedelta
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required
from extensions import db
from models import Funcionario, BancoHorasSaldo, AlocacaoDiaria
from services.banco_horas_service import calcular_saldo, salvar_saldos, get_config, set_config

financeiro_bp = Blueprint('financeiro', __name__)


@financeiro_bp.route('/banco-horas')
@login_required
def banco_horas():
    func_id = request.args.get('funcionario_id')
    data_inicio_str = request.args.get('data_inicio', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    data_fim_str = request.args.get('data_fim', date.today().strftime('%Y-%m-%d'))

    funcionarios = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()
    saldos = []
    funcionario_sel = None

    if func_id:
        funcionario_sel = Funcionario.query.get(func_id)
        d_ini = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        d_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        saldos = calcular_saldo(func_id, d_ini, d_fim)

    return render_template('banco_horas.html',
                           funcionarios=funcionarios,
                           funcionario_sel=funcionario_sel,
                           saldos=saldos,
                           data_inicio=data_inicio_str,
                           data_fim=data_fim_str)


@financeiro_bp.route('/banco-horas/alertas')
@login_required
def banco_horas_alertas():
    limite_dias = int(get_config('banco_horas_limite_dias', 30))
    data_limite = date.today() - timedelta(days=limite_dias)

    # Funcionários com saldo positivo antigo (horas a vencer)
    alertas = (
        db.session.query(
            BancoHorasSaldo.funcionario_id,
            Funcionario.nome,
            db.func.sum(BancoHorasSaldo.saldo_dia).label('saldo_total')
        )
        .join(Funcionario, Funcionario.id == BancoHorasSaldo.funcionario_id)
        .filter(BancoHorasSaldo.data <= data_limite, BancoHorasSaldo.saldo_dia > 0)
        .group_by(BancoHorasSaldo.funcionario_id, Funcionario.nome)
        .having(db.func.sum(BancoHorasSaldo.saldo_dia) > 0)
        .all()
    )

    return render_template('banco_horas_alertas.html', alertas=alertas, limite_dias=limite_dias)


@financeiro_bp.route('/config/banco-horas', methods=['GET', 'POST'])
@login_required
def config_banco_horas():
    if request.method == 'POST':
        set_config('banco_horas_data_fechamento', request.form.get('data_fechamento', '5'))
        set_config('banco_horas_limite_he_diario', request.form.get('limite_he_diario', '2'))
        set_config('banco_horas_limite_dias', request.form.get('limite_dias', '30'))
        set_config('banco_horas_valor_hora', request.form.get('valor_hora', '0'))
        flash('Configurações salvas!', 'success')
        return redirect(url_for('financeiro.config_banco_horas'))

    return render_template('config_banco_horas.html',
                           data_fechamento=get_config('banco_horas_data_fechamento', '5'),
                           limite_he=get_config('banco_horas_limite_he_diario', '2'),
                           limite_dias=get_config('banco_horas_limite_dias', '30'),
                           valor_hora=get_config('banco_horas_valor_hora', '0'))


@financeiro_bp.route('/banco-horas/exportar')
@login_required
def exportar_banco_horas():
    """RF3.2 – Exporta saldo de banco de horas em Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    func_id = request.args.get('funcionario_id')
    data_inicio_str = request.args.get('data_inicio', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    data_fim_str = request.args.get('data_fim', date.today().strftime('%Y-%m-%d'))

    func = Funcionario.query.get_or_404(func_id)
    d_ini = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
    d_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
    saldos = calcular_saldo(func_id, d_ini, d_fim)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Banco de Horas'

    # Cabeçalho
    header_fill = PatternFill('solid', fgColor='1E293B')
    header_font = Font(bold=True, color='38BDF8')
    headers = ['Data', 'Previsto (h)', 'Realizado (h)', 'Saldo Dia (h)', 'Saldo Acumulado (h)']
    ws.append(['Banco de Horas – ' + func.nome])
    ws.append([f'Período: {data_inicio_str} a {data_fim_str}'])
    ws.append([])
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Dados
    for s in saldos:
        ws.append([
            s['data'].strftime('%d/%m/%Y'),
            round(s['previsto'], 2),
            round(s['realizado'], 2),
            round(s['saldo_dia'], 2),
            round(s['saldo_acumulado'], 2),
        ])

    # Largura das colunas
    for i, width in enumerate([14, 14, 14, 14, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'banco_horas_{func.nome.replace(" ", "_")}_{data_inicio_str}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@financeiro_bp.route('/api/recalcular-banco-horas', methods=['POST'])
@login_required
def recalcular_banco_horas():
    """Persiste os saldos de banco de horas no banco para um funcionário ou todos."""
    payload = request.get_json(force=True, silent=True) or {}
    func_id = payload.get('funcionario_id')
    data_inicio_str = payload.get('data_inicio', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    data_fim_str = payload.get('data_fim', date.today().strftime('%Y-%m-%d'))
    d_ini = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
    d_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()

    if func_id:
        func = Funcionario.query.get_or_404(func_id)
        salvar_saldos(func.id, d_ini, d_fim)
        return jsonify({'ok': True, 'message': f'Saldos salvos para {func.nome}'})

    # Recalcular todos com alocações no período
    ids = {a.funcionario_id for a in
           AlocacaoDiaria.query.filter(AlocacaoDiaria.data >= d_ini).all()}
    erros = 0
    for fid in ids:
        try:
            salvar_saldos(fid, d_ini, d_fim)
        except Exception:
            erros += 1
    return jsonify({'ok': True, 'message': f'{len(ids)} funcionários recalculados ({erros} erros)'})


@financeiro_bp.route('/financeiro')
@login_required
def dashboard_financeiro():
    hoje = date.today()
    mes_inicio = hoje.replace(day=1)
    mes_anterior_fim = mes_inicio - timedelta(days=1)
    mes_anterior_inicio = mes_anterior_fim.replace(day=1)

    valor_hora = float(get_config('banco_horas_valor_hora', '0'))

    # Horas extras mês atual (saldo positivo)
    he_mes = db.session.query(
        db.func.sum(BancoHorasSaldo.saldo_dia)
    ).filter(
        BancoHorasSaldo.data >= mes_inicio,
        BancoHorasSaldo.data <= hoje,
        BancoHorasSaldo.saldo_dia > 0,
    ).scalar() or 0

    he_mes_anterior = db.session.query(
        db.func.sum(BancoHorasSaldo.saldo_dia)
    ).filter(
        BancoHorasSaldo.data >= mes_anterior_inicio,
        BancoHorasSaldo.data <= mes_anterior_fim,
        BancoHorasSaldo.saldo_dia > 0,
    ).scalar() or 0

    custo_he = float(he_mes) * valor_hora

    # Top 5 funcionários com mais horas extras
    top_he = (
        db.session.query(
            BancoHorasSaldo.funcionario_id,
            Funcionario.nome,
            db.func.sum(BancoHorasSaldo.saldo_dia).label('total_he')
        )
        .join(Funcionario, Funcionario.id == BancoHorasSaldo.funcionario_id)
        .filter(BancoHorasSaldo.data >= mes_inicio, BancoHorasSaldo.saldo_dia > 0)
        .group_by(BancoHorasSaldo.funcionario_id, Funcionario.nome)
        .order_by(db.text('total_he DESC'))
        .limit(5)
        .all()
    )

    return render_template('financeiro.html',
                           he_mes=round(float(he_mes), 1),
                           he_mes_anterior=round(float(he_mes_anterior), 1),
                           custo_he=round(custo_he, 2),
                           valor_hora=valor_hora,
                           top_he=top_he)


@financeiro_bp.route('/api/simular-escala')
@login_required
def simular_escala():
    """Simula impacto financeiro de alterar turno de um funcionário."""
    func_id = request.args.get('funcionario_id')
    turno_novo_horas = float(request.args.get('horas_novo', 0))
    turno_atual_horas = float(request.args.get('horas_atual', 0))
    valor_hora = float(get_config('banco_horas_valor_hora', '0'))
    dias_mes = 22  # dias úteis estimados

    delta_horas = (turno_novo_horas - turno_atual_horas) * dias_mes
    delta_custo = delta_horas * valor_hora

    return jsonify({
        'delta_horas': round(delta_horas, 1),
        'delta_custo': round(delta_custo, 2),
        'valor_hora': valor_hora,
    })
